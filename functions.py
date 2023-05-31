# functions for ./honeyz_proj/contact_finder

from openpyxl import load_workbook

path = './Customer List 05.27.xlsx'
workbook = load_workbook(filename = path)
current_sheet = workbook['Renamed Sheet']
row_counter = current_sheet.max_row
column_counter = current_sheet.max_column
print(f'There are {row_counter} rows in this sheet')
print(f'There are {column_counter} columns in this sheet.')
print('\n')
# range of sheet is A1:I1236
# def value_getter():
#     value_list = []
#     for value in current_sheet.iter_rows(
#             min_row=2,
#             max_row=10,
#             min_col=1,
#             max_col=10,
#             values_only=True):
#         # print(value)
#         # for each in value[0]:
#         #     print(type(each))
#         #     print(each)
#         #     num_check = each.isdigit()
#         #     print(bool(num_check))
#         # print(value[1])
#         value_list.append(value)
#     return value_list

def value_getter2():
    value_list = []
    for row in current_sheet.iter_rows(
            min_row=2,
            max_row=10,
            min_col=1,
            max_col=10,
            values_only=True):
        # for cell in row:
        #     if cell.value == "ABC":
        #         print('match found')
        value_list.append(row)
    return value_list

# for value in current_sheet.iter_rows(
#         min_row=2,
#         max_row=10,
#         min_col=1,
#         max_col=10,
#         values_only=True):
#     print(value)

x = value_getter2()
# print(x)
for each in x:
    for each_cell in each:
        print(f'{each_cell} is the {type(each_cell)} type.')
        if type(each_cell) is None:
            print('each cell is a none type. Converting to empty int')
            # each_cell = 0
            # print(type(each_cell))
            # each_cell = 1
            print(each_cell)
        elif type(each_cell) is int:
            print('cell is Int Type')
        print(each_cell)

# data = [current_sheet.cell(row=2,column=each).value for each in range(1,10)]
# print(data)
# print(data[6])
# print(type(data[6]))
#
# current_sheet[5:10]
# print(current_sheet)


# make for loop to go through each row
# def looper():
#     counter = 2
#     data_container = []
#     for each in data:
#         temp_holder =

# below finds val with criteria
# for each in data[4:]:
#     if each == int or each == None:
#         print("empty value")

# https://www.geeksforgeeks.org/get-values-of-all-rows-in-a-particular-column-in-openpyxl-python/


# def get_row_vals(crt_sht):
#     # names = []
#     row_container = []
#     appended_rows = []
#     for row in crt_sht:
#         p_name = row[0].value
#         row_container.append(p_name)
#         acct_num = row[1].value
#         row_container.append(acct_num)
#         prov = row[2].value
#         row_container.append(prov)
#         add = row[3].value
#         row_container.append(add)
#         city = row[4].value
#         row_container.append(city)
#         site_email1 = row[5].value
#         row_container.append(site_email1)
#         site_email2 = row[6].value
#         row_container.append(site_email2)
#         cmm_email = row[7].value
#         row_container.append(cmm_email)
#         gmid_email = row[8].value
#         row_container.append(gmid_email)
#         # names.append(p_name)
#         appended_rows.append(row_container)
#     # print(appended_rows)
#     # print(len(appended_rows))
#     # print('\n')
#     # print(appended_rows[5])
#     # print(row_container)
#     # for each in row_container:
#     #     print(each)
#     # print(appended_rows)
#     return appended_rows

# original working for loop to get row vals
def get_row_vals(crt_sht):
    names = []
    row_container = []
    for row in crt_sht:
        name = row[0].value
        row_container.append(name)
        names.append(name)
    return names



def print_out(list_of):
    for each in list_of:
        print(each)


def column_out(from_sheet_reader):
    with open("./raw_out.py", "w") as output:
        rows2str = str(from_sheet_reader)
        out_var = f'row_out = {rows2str}'
        output.write(out_var)


# get data from sheet
# perf_func = get_row_vals(current_sheet)
# print(perf_func[0])
# show_data()

# print out each value on sep line
# print_out(perf_func)

# save data to file as py
# column_out(perf_func)
